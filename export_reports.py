"""
Real Estate Report Exporter
============================
Connects to SQL Server, runs all 3 stored procedures,
and exports every result set into one Excel file —
one sheet per report section.

REQUIREMENTS (run once in Command Prompt):
    pip install pyodbc pandas openpyxl

EDIT THE CONFIG SECTION BELOW before running.
"""

import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter
import datetime
import os

# ============================================================
# CONFIG — edit these to match your SQL Server
# ============================================================
SERVER   = r'DESKTOP-9KT1UVF\SQLEXPRESS'
DATABASE = 'RealEstateDB'
# Windows Authentication (most common for local SQL Server)
CONN_STR = (
    f'DRIVER={{ODBC Driver 17 for SQL Server}};'
    f'SERVER={SERVER};'
    f'DATABASE={DATABASE};'
    f'Trusted_Connection=yes;'
)
# If using SQL login instead, comment above and uncomment below:
# USERNAME = 'sa'
# PASSWORD = ''
# CONN_STR = (
#     f'DRIVER={{ODBC Driver 17 for SQL Server}};'
#     f'SERVER={SERVER};DATABASE={DATABASE};'
#     f'UID={USERNAME};PWD={PASSWORD};'
# )

OUTPUT_FILE = r'C:\Users\USER\Desktop\RealEstate_FullReport.xlsx'

# ============================================================
# REPORT DEFINITIONS
# Each entry: (sheet_name, sql_to_execute)
# Multi-result-set procedures are handled automatically
# ============================================================
REPORTS = [
    # Market deep-dives
    ('Market_33993_2021',  "EXEC dbo.sp_GetMarketReport_Live @ZipCode='33993', @Year=2021"),
    ('Market_28906_2021',  "EXEC dbo.sp_GetMarketReport_Live @ZipCode='28906', @Year=2021"),
    ('Market_33160_2021',  "EXEC dbo.sp_GetMarketReport_Live @ZipCode='33160', @Year=2021"),

    # Agent scorecards
    ('Agents_NorthCarolina',"EXEC dbo.sp_AgentReport_Live @State='North Carolina', @MinSales=5"),
    ('Agents_Florida',      "EXEC dbo.sp_AgentReport_Live @State='Florida',        @MinSales=5"),
    ('Agents_Georgia',      "EXEC dbo.sp_AgentReport_Live @State='Georgia',        @MinSales=5"),

    # Investment index
    ('Investment_Index',   "EXEC dbo.sp_InvestmentIndex_Live @TopN=10, @MinListings=50"),

    # Bonus analytical queries
    ('A1_PricePerSqFt', """
        SELECT city, property_type, COUNT(*) AS TotalRecords,
            FORMAT(AVG(sale_price),'C0') AS AvgSalePrice,
            FORMAT(AVG(price_per_sqft),'C2') AS AvgPricePerSqFt,
            FORMAT(MIN(price_per_sqft),'C2') AS MinPricePerSqFt,
            FORMAT(MAX(price_per_sqft),'C2') AS MaxPricePerSqFt
        FROM dbo.vw_realtor_clean
        WHERE price_per_sqft BETWEEN 10 AND 5000
        GROUP BY city, property_type HAVING COUNT(*) >= 5
        ORDER BY city, AVG(price_per_sqft) DESC
    """),
    ('B2_FastestMarkets', """
        SELECT TOP 20 zip_code, city, state, COUNT(*) AS ListingCount,
            FORMAT(AVG(CAST(days_on_market AS FLOAT)),'N1') AS AvgDaysOnMarket,
            FORMAT(AVG(sale_price),'C0') AS AvgSalePrice
        FROM dbo.vw_realtor_clean
        WHERE days_on_market BETWEEN 0 AND 365 AND sale_price IS NOT NULL
        GROUP BY zip_code, city, state HAVING COUNT(*) >= 10
        ORDER BY AVG(CAST(days_on_market AS FLOAT)) ASC
    """),
    ('C1_AgentRankings', """
        SELECT TOP 50
            brokered_by AS AgentBroker, state, years_experience,
            COUNT(*) AS TxCount,
            FORMAT(SUM(sale_price),'C0') AS TotalVolume,
            FORMAT(AVG(sale_price),'C0') AS AvgSalePrice,
            FORMAT(AVG(CAST(days_on_market AS FLOAT)),'N1') AS AvgDOM,
            RANK() OVER (ORDER BY SUM(sale_price) DESC) AS VolumeRank
        FROM dbo.vw_realtor_clean
        WHERE brokered_by IS NOT NULL
        GROUP BY brokered_by, state, years_experience
        HAVING COUNT(*) >= 3
        ORDER BY SUM(sale_price) DESC
    """),
    ('D1_InvestmentSignals', """
        SELECT TOP 20 zip_code, city, state,
            COUNT(*) AS TotalListings,
            FORMAT(AVG(sale_price),'C0') AS AvgSalePrice,
            FORMAT(AVG(price_per_sqft),'C2') AS AvgPricePerSqFt,
            FORMAT(AVG(CAST(days_on_market AS FLOAT)),'N1') AS AvgDaysOnMarket,
            FORMAT(AVG(sale_to_list_ratio),'P2') AS AvgSaleToList,
            FORMAT(
                CAST(SUM(CASE WHEN status IN ('sold','s') THEN 1 ELSE 0 END) AS FLOAT)
                / NULLIF(COUNT(*),0),'P1') AS SoldRate
        FROM dbo.vw_realtor_clean
        WHERE sale_price IS NOT NULL
        GROUP BY zip_code, city, state HAVING COUNT(*) >= 20
        ORDER BY AVG(sale_price) DESC
    """),
]

# ============================================================
# STYLES
# ============================================================
NAVY   = '1B3A6B'
BLUE   = '2E6DA4'
LTBLUE = 'D6E4F0'
GRAY   = 'F4F6F8'
WHITE  = 'FFFFFF'
GREEN  = '1E6B3A'
LTGREEN= 'D6F0DF'

def header_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def style_sheet(ws, df, title, section_color=NAVY):
    """Apply professional formatting to a worksheet."""
    # Title row
    ws.insert_rows(1)
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(name='Arial', size=14, bold=True, color=WHITE)
    title_cell.fill      = header_fill(section_color)
    title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(df.columns))

    # Timestamp row
    ts_cell = ws.cell(row=2, column=1,
        value=f'Generated: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}  |  Source: RealEstateDB.dbo.realtor_data')
    ts_cell.font      = Font(name='Arial', size=9, italic=True, color='666666')
    ts_cell.alignment = Alignment(horizontal='left', indent=1)
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=len(df.columns))

    # Header row (row 3)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = Font(name='Arial', size=10, bold=True, color=WHITE)
        cell.fill      = header_fill(NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[3].height = 22

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=4):
        fill = header_fill(LTBLUE) if row_idx % 2 == 0 else header_fill(WHITE)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name='Arial', size=10)
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx-1].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Freeze panes below header
    ws.freeze_panes = 'A4'

# ============================================================
# MAIN EXPORT
# ============================================================
def run_query_all_results(cursor, sql):
    """Execute SQL and return ALL result sets as a list of DataFrames."""
    cursor.execute(sql)
    results = []
    while True:
        if cursor.description:
            cols = [d[0] for d in cursor.description]
            rows = cursor.fetchall()
            df   = pd.DataFrame.from_records(rows, columns=cols)
            results.append(df)
        if not cursor.nextset():
            break
    return results

def main():
    print(f'\nConnecting to {SERVER} / {DATABASE}...')
    try:
        conn   = pyodbc.connect(CONN_STR, timeout=30)
        cursor = conn.cursor()
        print('Connected successfully.\n')
    except Exception as e:
        print(f'Connection failed: {e}')
        print('\nTroubleshooting:')
        print('  1. Check SERVER name in CONFIG section')
        print('  2. Make sure SQL Server is running')
        print('  3. Try SERVER = "." or "localhost" or "DESKTOP-XXXX\\SQLEXPRESS"')
        return

    # Create workbook
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_count = 0
    section_colors = [NAVY, BLUE, GREEN, NAVY, BLUE, GREEN, NAVY, BLUE, GREEN, GREEN, NAVY, GREEN]

    for report_idx, (sheet_name, sql) in enumerate(REPORTS):
        print(f'Running: {sheet_name}...')
        try:
            results = run_query_all_results(cursor, sql)

            if not results or all(df.empty for df in results):
                print(f'  → No data returned for {sheet_name}')
                continue

            color = section_colors[report_idx % len(section_colors)]

            # Multi-result-set procedures get sub-sheets
            if len(results) > 1:
                section_labels = ['KPI Summary', 'By Property Type',
                                  'By Financing Type', 'Transactions']
                for sub_idx, df in enumerate(results):
                    if df.empty:
                        continue
                    label = section_labels[sub_idx] if sub_idx < len(section_labels) \
                            else f'Result {sub_idx+1}'
                    full_name = f'{sheet_name[:18]}_{label[:10]}'[:31]
                    ws = wb.create_sheet(title=full_name)
                    style_sheet(ws, df,
                        title=f'{sheet_name.replace("_"," ")} — {label}',
                        section_color=color)
                    sheet_count += 1
                    print(f'  → Sheet: {full_name} ({len(df)} rows)')
            else:
                df = results[0]
                ws = wb.create_sheet(title=sheet_name[:31])
                style_sheet(ws, df,
                    title=sheet_name.replace('_', ' '),
                    section_color=color)
                sheet_count += 1
                print(f'  → Sheet: {sheet_name} ({len(df)} rows)')

        except Exception as e:
            print(f'  → Error on {sheet_name}: {e}')
            continue

    # Summary sheet at the front
    ws_summary = wb.create_sheet(title='README', index=0)
    summary_data = [
        ['Real Estate Analytics Report'],
        [f'Generated: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'],
        [''],
        ['Source',   'RealEstateDB.dbo.realtor_data'],
        ['Records',  '1,048,575'],
        ['Server',   SERVER],
        [''],
        ['Sheet', 'Contents'],
        ['Market_33993_2021', 'Zip 33993 market deep-dive — KPI, property type, financing, transactions'],
        ['Market_28906_2021', 'Zip 28906 market deep-dive'],
        ['Market_33160_2021', 'Zip 33160 market deep-dive'],
        ['Agents_NorthCarolina', 'Agent scorecard — North Carolina'],
        ['Agents_Florida',       'Agent scorecard — Florida'],
        ['Agents_Georgia',       'Agent scorecard — Georgia'],
        ['Investment_Index',     'Top 10 zip codes by weighted investment score'],
        ['A1_PricePerSqFt',      'Avg price per sqft by city and property type'],
        ['B2_FastestMarkets',    'Top 20 fastest-moving zip codes by days on market'],
        ['C1_AgentRankings',     'Top 50 agents by total sales volume'],
        ['D1_InvestmentSignals', 'Top 20 zip codes by avg sale price'],
    ]
    for r_idx, row in enumerate(summary_data, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(name='Arial', size=16, bold=True, color=NAVY)
            elif r_idx == 8:
                cell.font = Font(name='Arial', size=11, bold=True, color=WHITE)
                cell.fill = header_fill(NAVY)
            elif r_idx > 8:
                cell.font = Font(name='Arial', size=10,
                    color=NAVY if c_idx == 1 else '333333')
                cell.fill = header_fill(LTBLUE if r_idx % 2 == 0 else WHITE)
            else:
                cell.font = Font(name='Arial', size=10, color='555555')
    ws_summary.column_dimensions['A'].width = 28
    ws_summary.column_dimensions['B'].width = 55

    # Save
    wb.save(OUTPUT_FILE)
    conn.close()
    print(f'\n✓ Done! {sheet_count} sheets exported to:')
    print(f'  {OUTPUT_FILE}')
    print('\nOpen the file in Excel — each report is a separate formatted sheet.')

if __name__ == '__main__':
    main()